import streamlit as st
import openpyxl as px
from io import BytesIO

st.title('EPMAデータをコピペ！')
st.write('<p style="font-size: 20px;">EPMAから取得したエクセルデータを補正用エクセルファイルに貼り付けます</p>', unsafe_allow_html=True)

with st.form(key="fileandname"):
    #ファイルアップローダー
    EPMAfile = st.file_uploader("☟EPMAのエクセルデータをアップロードしてください",type="xlsx")
    #測定間隔
    micron = st.number_input("☟測定の間隔は何µmですか？",0,30,1)
    #完成ファイルの名前
    filename = st.text_input("☟処理後のファイル名を入力してください","")
    #ボタン
    submit_btn = st.form_submit_button('GO!')


if submit_btn:
    template = px.load_workbook("./template/Ca-Sr-Mg-Sver3.xlsx")
    #EPMA結果（エクセルファイル）を読み込み
    result = px.load_workbook(EPMAfile)
    N = len(result.sheetnames)#シートの数をカウント
    pages = 0 #テンプレートの何シート目を使うのかを数えるカウンター
    #EPMA結果のページ分だけ...
    for i in range(N):
        resultn = result.worksheets[i]#コピー元シート
        endpoint = resultn.max_row-6
        endpointlist = [3]
        #↓個体が切り替わる直前の行をendpointlistに記録
        while True:
            if endpoint <= 3:
                break
            else:
                endpointlist.insert(1,endpoint)
                points = resultn.cell(row=endpoint, column=2).value.split()[-1]
                endpoint -= int(points)


        for j in range(len(endpointlist)-1):#i番目のシートにある個体全てを扱うためのfor文
            pages += 1
            templaten = template.worksheets[pages-1]
            idnum = resultn.cell(row=endpointlist[j]+1,column=2).value
            id = str(idnum.split()[0])
            #templaten.title = id #シート名変更をするとリンク？がおかしくなるのでエスケープ
            for k in range(endpointlist[j]+1,endpointlist[j+1]+1):#i番目のシートk番目の個体について各行をコピーするためのfor文
                #個体番号を入力
                templaten.cell(row=k-endpointlist[j]+2,column=1,value=id)
                #測定ポイント番号を入力
                num = (int(resultn.cell(row=k,column=2).value.split()[-1])-1) * int(micron)
                templaten.cell(row=k-endpointlist[j]+2,column=2,value=num)
                #CaOをコピペ
                copyCaO = resultn.cell(row=k,column=4).value
                templaten.cell(row=k-endpointlist[j]+2,column=3,value=copyCaO)
                #Total(mass)をコピペ
                copyTotalMass = resultn.cell(row = k, column = 8).value
                templaten.cell(row=k-endpointlist[j]+2, column=4, value=copyTotalMass)
                #SrOからTotal(Norm%)までをコピペ
                for l in range(9,15):
                    copy = resultn.cell(row = k, column = l).value
                    templaten.cell(row = k-endpointlist[j]+2, column = l-4, value = copy)

    for n in range(pages+1,101):
        template.remove(template.worksheets[-1])
    #template.to_excel(buf := BytesIO(), index=False)
    output = BytesIO()
    template.save(output)
    output.seek(0)
    st.download_button(
        "Download",
        output,
        f'{filename}.xlsx',
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

st.write('<p style="font-size: 15px;">※Parameter:ALLのエクセルファイルのみ処理可能です<br>※各シート内のコメントが同じ測定点をまとめてひとつのシートにまとめます<br>※処理中はページ右上に"Running"が表示され、処理が終わると"Download"ボタンが出現します</p>', unsafe_allow_html=True)
